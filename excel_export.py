# Contenu de excel_export.py
import io
import pandas as pd
import openpyxl
import json # Importé pour la sauvegarde
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# La signature est correcte, elle accepte les deux types de données
def create_styled_excel(project_info_dict, df_data_edited, t_tb_raw_val, save_data_dict=None):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    wb = writer.book
    # Crée la feuille de débit visible
    ws = wb.create_sheet(title="Feuille de Débit", index=0)
    
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # --- REMIS : TOUT LE CODE DE STYLES ET D'ÉCRITURE DE LA FEUILLE DE DÉBIT ---
    
    # Styles
    bold_font = Font(bold=True)
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Ligne 2: Client
    ws['B2'] = "Client :"
    ws['B2'].font = bold_font
    ws['C2'] = project_info_dict['client']
    ws.merge_cells('C2:F2')

    # Ligne 3: Réf Chantier
    ws['B3'] = "Réf Chantier :"
    ws['B3'].font = bold_font
    ws['C3'] = project_info_dict['ref_chantier']
    ws.merge_cells('C3:F3')

    # Ligne 4: Téléphone
    ws['B4'] = "Téléphone / Mail :"
    ws['B4'].font = bold_font
    ws['C4'] = project_info_dict['telephone']
    ws.merge_cells('C4:F4')

    # Ligne 2: Date
    ws['G2'] = "Date :"
    ws['G2'].font = bold_font
    ws['H2'] = project_info_dict['date']

    # Ligne 7: DEVIS/COMMANDE
    ws['B7'] = "DEVIS"
    ws['B7'].fill = pink_fill
    ws['C7'] = "COMMANDE"
    ws['C7'].fill = pink_fill
    
    ws['E7'] = "Date souhaitée"
    ws['E7'].fill = pink_fill
    ws['F7'] = project_info_dict['date_souhaitee']
    ws['F7'].alignment = Alignment(horizontal='center')
    ws.merge_cells('F7:G7')
    ws['F7'].fill = pink_fill
    ws['G7'].fill = pink_fill


    # Ligne 9: Panneau/Décor
    ws['B9'] = "Panneau / Décor :"
    ws['B9'].font = bold_font
    ws['B9'].fill = green_fill
    ws['C9'] = project_info_dict['panneau_decor']
    ws.merge_cells('C9:F9')
    for col in ['C','D','E','F']: ws[f'{col}9'].fill = green_fill

    ws['G9'] = "Epaisseur :"
    ws['G9'].font = bold_font
    ws['G9'].fill = green_fill
    ws['H9'] = t_tb_raw_val # Récupère l'épaisseur
    ws['H9'].fill = green_fill

    # Ligne 10: Chant
    ws['B10'] = "Chant :"
    ws['B10'].font = bold_font
    ws['B10'].fill = green_fill
    ws['D10'] = project_info_dict['chant_mm']
    ws['D10'].fill = green_fill
    ws['F10'] = "Décor :"
    ws['F10'].font = bold_font
    ws['F10'].fill = green_fill
    ws['G10'] = project_info_dict['decor_chant']
    ws.merge_cells('G10:H10')
    ws['G10'].fill = green_fill
    ws['H10'].fill = green_fill

    # --- Écriture du DataFrame ---
    
    # Convertir les booléens en OUI/NON pour l'export
    df_export = df_data_edited.copy()
    
    for col in ["Chant Avant", "Chant Arrière", "Chant Gauche", "Chant Droit"]:
        if col in df_export.columns:
            # Gère le cas où la colonne n'existe pas (si le df est vide)
            df_export[col] = df_export[col].map({True: 'OUI', False: 'NON', 'nan': 'NON'})

    # Ligne de titre pour le tableau
    header_row = 13
    ws.cell(row=header_row, column=1, value="N°").font = bold_font
    ws.cell(row=header_row, column=2, value="Référence Pièce").font = bold_font
    ws.cell(row=header_row, column=3, value="Qté").font = bold_font
    ws.cell(row=header_row, column=4, value="Longueur en mm").font = bold_font
    ws.cell(row=header_row, column=5, value="Chant").font = bold_font
    ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=6)
    ws.cell(row=header_row+1, column=5, value="Avant").font = bold_font
    ws.cell(row=header_row+1, column=6, value="Arrière").font = bold_font
    ws.cell(row=header_row, column=7, value="Largeur en mm").font = bold_font
    ws.cell(row=header_row, column=8, value="Chant").font = bold_font
    ws.merge_cells(start_row=header_row, start_column=8, end_row=header_row, end_column=9)
    ws.cell(row=header_row+1, column=8, value="Gauche").font = bold_font
    ws.cell(row=header_row+1, column=9, value="Droit").font = bold_font
    ws.cell(row=header_row, column=10, value="Usinage (*)").font = bold_font
    
    # Centrer les en-têtes
    for r in range(header_row, header_row + 2):
        for c in range(1, 11):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=c).fill = grey_fill
            ws.cell(row=r, column=c).border = thin_border

    # Écrire les données du DataFrame
    start_data_row = header_row + 2
    
    # Utilise .itertuples() pour écrire les données
    # Index 0: Lettre, 1: Réf Pièce, 2: Qté, 3: L, 4: ChA, 5: ChP, 6: l, 7: ChG, 8: ChD, 9: Usinage
    # Doit correspondre à l'ordre dans le DataFrame 'df_global'
    for r_idx, row in enumerate(df_export.itertuples(index=False), start=start_data_row):
        ws.cell(row=r_idx, column=1, value=row[0]) # Lettre
        ws.cell(row=r_idx, column=2, value=row[1]) # Référence Pièce
        ws.cell(row=r_idx, column=3, value=row[2]) # Qté
        ws.cell(row=r_idx, column=4, value=row[3]) # Longueur (mm)
        ws.cell(row=r_idx, column=5, value=row[4]) # Chant Avant
        ws.cell(row=r_idx, column=6, value=row[5]) # Chant Arrière
        ws.cell(row=r_idx, column=7, value=row[6]) # Largeur (mm)
        ws.cell(row=r_idx, column=8, value=row[7]) # Chant Gauche
        ws.cell(row=r_idx, column=9, value=row[8]) # Chant Droit
        ws.cell(row=r_idx, column=10, value=row[9]) # Usinage
        
        # Appliquer les bordures aux données
        for c in range(1, 11):
            ws.cell(row=r_idx, column=c).border = thin_border
            if c in [3, 5, 6, 8, 9]:
                ws.cell(row=r_idx, column=c).alignment = Alignment(horizontal='center')

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['J'].width = 25
    
    # --- FIN DU CODE DE LA FEUILLE DE DÉBIT ---


    # --- AJOUT DES DONNÉES DE SAUVEGARDE (Inchangé) ---
    if save_data_dict:
        try:
            ws_data = wb.create_sheet(title="SaveData")
            json_string = json.dumps(save_data_dict, indent=2)
            ws_data['A1'] = json_string
            ws_data.sheet_state = 'hidden' # Cache la feuille
        except Exception as e:
            print(f"Erreur lors de l'écriture des données de sauvegarde : {e}")
            pass
    # --- FIN DU BLOC DE SAUVEGARDE ---
    
    writer.close() 
    return output.getvalue()